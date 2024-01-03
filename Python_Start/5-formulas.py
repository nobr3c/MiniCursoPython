from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 1 - Lê pasta de trabalho e planilha 
wb = load_workbook("data/barchart.xlsx")
sheet = wb["Relatorio"]

# 2 - referências das linhas e Colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# # 3 - Incluindo Fórmula
# sheet["B6"] = "=SUM(B2:B5)"
# sheet["B6"].style = "Currency"
# wb.save("data/test.xlsx")

for i in range(min_column+1, max_column+1): #inicia a partir da coluna 2 e da linha 2
    letter = get_column_letter(i)
    # print(letter)
    sheet[f"{letter}{max_row+1}"] = f"=SUM({letter}{min_row+1}:{letter}{max_row})"
    sheet[f"{letter}{max_row+1}"].style = "Currency"
    wb.save("data/testAut.xlsx")